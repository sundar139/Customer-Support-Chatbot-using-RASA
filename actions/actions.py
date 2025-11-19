from typing import Any, Text, Dict, List
from rasa_sdk import Action, Tracker, FormValidationAction
from rasa_sdk.executor import CollectingDispatcher
from rasa_sdk.types import DomainDict
from rasa_sdk.events import SlotSet, FollowupAction
import json
from pathlib import Path
import os
from dotenv import load_dotenv
import uuid
from datetime import datetime
from typing import Optional

# Excel logging
try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

load_dotenv()


class OrderDatabase:
    def __init__(self):
        db_path = os.getenv('ORDER_DATABASE_PATH', './dataset/orders.json')
        self.db_path = Path(db_path)
        self.orders = self._load_orders()
    
    def _load_orders(self):
        if self.db_path.exists():
            with open(self.db_path, 'r') as f:
                return json.load(f)
        return {}
    
    def get_order(self, order_id: str):
        return self.orders.get(order_id)
    
    def mark_return(self, order_id: str, reason: str):
        if order_id in self.orders:
            self.orders[order_id]['return_requested'] = True
            self.orders[order_id]['return_reason'] = reason
            self._save_orders()
            return True
        return False
    
    def _save_orders(self):
        with open(self.db_path, 'w') as f:
            json.dump(self.orders, f, indent=2)


db = OrderDatabase()


def append_ticket_log(issue_id: str, summary: str, order_id: Optional[str], sender_id: Optional[str]):
    """Append a ticket entry to dataset/tickets.xlsx. Creates the workbook if missing."""
    try:
        if Workbook is None:
            return
        path = Path('./dataset/tickets.xlsx')
        if path.exists():
            wb = load_workbook(path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'tickets'
            ws.append(['timestamp', 'issue_id', 'order_id', 'sender_id', 'summary'])
        ws.append([
            datetime.utcnow().isoformat(timespec='seconds') + 'Z',
            issue_id,
            order_id or '',
            sender_id or '',
            summary,
        ])
        wb.save(path)
    except Exception:
        # Avoid crashing action on logging errors
        pass


class ActionCheckOrderStatus(Action):
    
    def name(self) -> Text:
        return "action_check_order_status"
    
    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        
        order_id = tracker.get_slot("order_id")
        
        if not order_id:
            dispatcher.utter_message(text="I need an order ID to check the status. Could you provide it?")
            return []
        
        order = db.get_order(order_id)
        
        if order:
            status = order.get('status', 'unknown')

            # Show only essential details; avoid explicit status strings
            if status == 'in_transit':
                expected_delivery = order.get('expected_delivery', 'soon')
                tracking = order.get('tracking_number', 'N/A')
                message = f"Expected delivery by {expected_delivery}. Tracking number: {tracking}."
            elif status == 'delivered':
                delivery_date = order.get('delivery_date', 'recently')
                message = f"Delivered on {delivery_date}."
            elif status == 'processing':
                expected_delivery = order.get('expected_delivery', 'soon')
                message = f"Expected delivery: {expected_delivery}."
            else:
                # Generic detail when status is unknown
                expected_delivery = order.get('expected_delivery')
                tracking = order.get('tracking_number')
                parts = []
                if expected_delivery:
                    parts.append(f"Expected delivery: {expected_delivery}")
                if tracking:
                    parts.append(f"Tracking number: {tracking}")
                message = ". ".join(parts) or "Order update available."

            items = order.get('items', [])
            if items:
                message += f"\nItems: {', '.join(items)}"
        
        else:
            message = f"I'm sorry, I couldn't find any order with ID {order_id}. Please check the order ID and try again."
        
        dispatcher.utter_message(text=message)
        
        return []


class ActionProcessReturn(Action):
    
    def name(self) -> Text:
        return "action_process_return"
    
    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        
        order_id = tracker.get_slot("order_id")
        return_reason = tracker.get_slot("return_reason")
        
        if not order_id:
            dispatcher.utter_message(text="I need an order ID to process the return.")
            return []
        
        order = db.get_order(order_id)
        
        if not order:
            dispatcher.utter_message(text=f"I couldn't find order {order_id}. Please verify the order ID.")
            return []
        
        success = db.mark_return(order_id, return_reason or "Not specified")
        
        if success:
            message = f"Your return request for order {order_id} has been successfully processed.\n\n"
            message += "You'll receive a confirmation email with return shipping instructions within 24 hours.\n"
            message += "Refunds are typically processed within 5-7 business days after we receive the returned item."
            
            if return_reason:
                message += f"\n\nReturn reason recorded: {return_reason}"
        else:
            message = f"There was an issue processing the return for order {order_id}. Please try again or contact support."
        
        dispatcher.utter_message(text=message)
        
        return [SlotSet("order_id", None), SlotSet("return_reason", None)]


class ValidateOrderStatusForm(FormValidationAction):
    
    def name(self) -> Text:
        return "validate_order_status_form"
    
    def validate_order_id(
        self,
        slot_value: Any,
        dispatcher: CollectingDispatcher,
        tracker: Tracker,
        domain: DomainDict,
    ) -> Dict[Text, Any]:
        
        if slot_value and len(str(slot_value)) == 5 and str(slot_value).isdigit():
            return {"order_id": slot_value}
        else:
            dispatcher.utter_message(text="Order ID should be a 5-digit number. Please try again.")
            return {"order_id": None}


class ValidateReturnForm(FormValidationAction):
    
    def name(self) -> Text:
        return "validate_return_form"
    
    def validate_order_id(
        self,
        slot_value: Any,
        dispatcher: CollectingDispatcher,
        tracker: Tracker,
        domain: DomainDict,
    ) -> Dict[Text, Any]:
        
        if slot_value and len(str(slot_value)) == 5 and str(slot_value).isdigit():
            order = db.get_order(str(slot_value))
            if order:
                return {"order_id": slot_value}
            else:
                dispatcher.utter_message(text=f"I couldn't find order {slot_value}. Please check the order ID.")
                return {"order_id": None}
        else:
            dispatcher.utter_message(text="Order ID should be a 5-digit number.")
            return {"order_id": None}
    
    def validate_return_reason(
        self,
        slot_value: Any,
        dispatcher: CollectingDispatcher,
        tracker: Tracker,
        domain: DomainDict,
    ) -> Dict[Text, Any]:
        
        if slot_value and len(str(slot_value)) > 0:
            return {"return_reason": slot_value}
        else:
            return {"return_reason": "Not specified"}


class ActionDefaultFallback(Action):
    
    def name(self) -> Text:
        return "action_default_fallback"
    
    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        
        latest_text = tracker.latest_message.get("text") or "Fallback triggered"
        order_id = tracker.get_slot("order_id")
        prior_issue = tracker.get_slot("issue_id")
        fallback_count = tracker.get_slot("fallback_count") or 0

        # If user said "yes" and we already have an order_id, route to status check
        latest_intent = (tracker.latest_message.get("intent") or {}).get("name")
        if latest_intent == "affirm" and order_id:
            return [FollowupAction("action_check_order_status"), SlotSet("fallback_count", 0)]

        # If we've already fallen back recently or a ticket exists, avoid spamming tickets
        if prior_issue or (isinstance(fallback_count, (int, float)) and fallback_count >= 1):
            dispatcher.utter_message(text=(
                "I'm not sure I understand. Can you rephrase your question?"
            ))
            return [SlotSet("fallback_count", (fallback_count or 0) + 1)]

        # Otherwise, generate a support ticket once
        rand = uuid.uuid4().hex[:6].upper()
        issue_id = f"ISSUE-{rand}"
        summary = f"Ticket for order {order_id}: {latest_text}" if order_id else latest_text

        # Log ticket to Excel and avoid exposing ticket id in chat
        try:
            append_ticket_log(issue_id, summary, order_id, tracker.sender_id)
        except Exception:
            pass

        dispatcher.utter_message(text=(
            "I've created a support ticket and captured your issue details. If you share your order ID, I can provide delivery information."
        ))

        return [SlotSet("issue_id", issue_id), SlotSet("problem_summary", summary), SlotSet("fallback_count", 1)]


class ActionStoreOrderId(Action):
    
    def name(self) -> Text:
        return "action_store_order_id"
    
    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        
        # Try to extract order_id entity from the latest user message
        order_id_entity = None
        for e in tracker.latest_message.get("entities", []):
            if e.get("entity") == "order_id" and e.get("value"):
                order_id_entity = e.get("value")
                break
        
        # Fallback to the existing slot if entity not present
        current_slot = tracker.get_slot("order_id")
        order_id = order_id_entity or current_slot
        
        if order_id:
            dispatcher.utter_message(text=f"Thanks! I have recorded your order ID: {order_id}. Would you like me to check its status?")
            # Reset fallback counter once we have a clear path
            return [SlotSet("order_id", order_id), SlotSet("fallback_count", 0)]
        else:
            dispatcher.utter_message(text="I couldn't detect an order ID. Please share it and I'll take it from there.")
            return []


class ActionCreateTicket(Action):

    def name(self) -> Text:
        return "action_create_ticket"

    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:

        # Build a brief summary from the latest user message
        latest_text = tracker.latest_message.get("text") or "Issue reported"

        # Generate a simple human-readable issue id
        rand = uuid.uuid4().hex[:6].upper()
        issue_id = f"ISSUE-{rand}"

        # Optionally attach order_id if present
        order_id = tracker.get_slot("order_id")
        if order_id:
            summary = f"Ticket for order {order_id}: {latest_text}"
        else:
            summary = latest_text

        # Log ticket to Excel and keep chat concise (no ticket id shown)
        try:
            append_ticket_log(issue_id, summary, order_id, tracker.sender_id)
        except Exception:
            pass

        dispatcher.utter_message(text=(
            "Thanks, I've created a support ticket and noted your issue. If you have an order ID, I can provide delivery details."
        ))

        # Store issue_id and a brief summary; also reset fallback counter
        return [SlotSet("issue_id", issue_id), SlotSet("problem_summary", summary), SlotSet("fallback_count", 0)]